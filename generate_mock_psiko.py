#!/usr/bin/env python3
"""
Create an anonymized copy of the psychotechnical tracker's source files for a
public portfolio.

Three source tables, masked independently (they share no people):
  - Kursiyer Listesi   : course trainees, keyed by Aday No
  - Psiko Sonuçları     : call results, keyed by Aday No (joins to the trainee list)
  - Havuz (x2)          : externally sourced people, no Aday No, keyed by date

Aday No is mapped consistently WITHIN each file so the trainee-list <-> results
join survives. Names, phone numbers, and ID-like columns are replaced. Dates are
left untouched: the segmentation DAX anchors on the fixed legal date 2021-06-30,
so shifting dates would break it, and a bare date next to a fake name identifies
no one. Columns that identify a person but never appear in the report
(T.C. Kimlik No, Sertifika No) are filled with obvious placeholders so they can
be deleted afterward without leaving a real value behind.

Set the file paths below, then run:  python generate_mock.py
"""

import os
import random

from openpyxl import Workbook, load_workbook

# ---- File paths (edit these) --------------------------------------------

TRAINEE_PATH  = r"candidates"
RESULTS_PATH  = r"result"
POOL_PATHS    = [
    r"C:psycho-2025",
    r"psycho-2026",
]
OUT_DIR       = r"C:\Users\dogan\OneDrive\Masaüstü\Power BI Project\mock"

SEED = 20260719  # fixed so re-runs produce the same anonymization

# ---- Column handling -----------------------------------------------------
# Exact header text as it appears in each file. Adjust if yours differ.

ADAY_NO   = "ADAY NO"
FIRST     = "ADI"
LAST      = "SOYADI"
FULLNAME  = "ADI-SOYADI"   # pool files store the name in one column
PHONE     = "TELEFON"

# Columns that identify a person but don't appear in the report. Filled with a
# placeholder here; delete them entirely after masking.
DROP_LATER = ["T.C KİMLİK NO", "SERTİFİKA NO"]

# --------------------------------------------------------------------------


def rng_phone(rng):
    """A 10-digit random phone number as a string."""
    return "".join(str(rng.randint(0, 9)) for _ in range(10))


def read_sheet(path):
    workbook = load_workbook(path)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    headers = list(rows[0])
    data = [list(r) for r in rows[1:]]
    return headers, data, worksheet.title


def write_sheet(path, headers, data, title):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    worksheet.append(headers)
    for row in data:
        worksheet.append(row)
    workbook.save(path)


def idx(headers, name):
    """Index of a column, or None if the file doesn't have it."""
    return headers.index(name) if name in headers else None


def mask_aday_no(headers, data, rng):
    """Replace Aday No with a consistent fake number, per file."""
    i = idx(headers, ADAY_NO)
    if i is None:
        return
    mapping = {}
    for row in data:
        raw = row[i]
        if raw in (None, ""):
            continue
        if raw not in mapping:
            mapping[raw] = len(mapping) + 1
        row[i] = mapping[raw]


def mask_names(headers, data, rng):
    """Replace names with 'Aday N', consistent within this file.

    Handles both split (ADI + SOYADI) and combined (ADI-SOYADI) layouts.
    """
    first_i = idx(headers, FIRST)
    last_i  = idx(headers, LAST)
    full_i  = idx(headers, FULLNAME)

    mapping = {}

    def fake_for(key):
        if key not in mapping:
            mapping[key] = f"Aday {len(mapping) + 1}"
        return mapping[key]

    for row in data:
        if full_i is not None:
            raw = row[full_i]
            if raw not in (None, ""):
                row[full_i] = fake_for(str(raw).strip().casefold())
        elif first_i is not None:
            key = f"{row[first_i]} {row[last_i] if last_i is not None else ''}".strip().casefold()
            fake = fake_for(key)
            row[first_i] = fake
            if last_i is not None:
                row[last_i] = ""   # whole name now lives in ADI


def mask_phone(headers, data, rng):
    i = idx(headers, PHONE)
    if i is None:
        return
    for row in data:
        if row[i] not in (None, ""):
            row[i] = rng_phone(rng)


def placeholder_drop_columns(headers, data):
    """Overwrite identify-but-unused columns with an obvious placeholder."""
    for name in DROP_LATER:
        i = idx(headers, name)
        if i is None:
            continue
        for row in data:
            if row[i] not in (None, ""):
                row[i] = "REMOVE_ME"


def process(path, rng, out_dir):
    headers, data, title = read_sheet(path)
    mask_aday_no(headers, data, rng)
    mask_names(headers, data, rng)
    mask_phone(headers, data, rng)
    placeholder_drop_columns(headers, data)
    out_path = os.path.join(out_dir, os.path.basename(path))
    write_sheet(out_path, headers, data, title)
    return out_path


def main():
    rng = random.Random(SEED)
    os.makedirs(OUT_DIR, exist_ok=True)

    outputs = []
    for path in [TRAINEE_PATH, RESULTS_PATH, *POOL_PATHS]:
        outputs.append(process(path, rng, OUT_DIR))

    print("Masked:")
    for path in outputs:
        print(f"  {path}")
    print(
        "\nDates left untouched (segmentation anchors on 2021-06-30).\n"
        "Columns marked REMOVE_ME (T.C. Kimlik No, Sertifika No) should be "
        "deleted in Power BI after loading.\n"
        "Then point the data sources at these files and refresh."
    )


if __name__ == "__main__":
    main()
